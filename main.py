import os
import re
import time
from datetime import datetime, timedelta
from typing import List

import openpyxl
import pytz
import firebase_admin
from firebase_admin import credentials, db
from openpyxl.styles import Alignment
from telegram import *
from telegram.ext import *
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import get_column_letter

cred = credentials.Certificate("kit-pro-f4b0d-firebase-adminsdk-mhzrf-8a07acab1c.json")
firebase_admin.initialize_app(cred, {
    "databaseURL": "https://kit-pro-f4b0d-default-rtdb.firebaseio.com/"
})

# Set up the Telegram bot

bot = Bot(token="6208523031:AAFfOb97T6Wml0pZUagE56A_MZDpCpUXZJk")


def start(update, context):
    message = update.message
    chat_id = message.chat_id
    bot.sendMessage(chat_id=chat_id,
                    text="Hi! I'm your Telegram bot. I'll collect messages and links from PoolSea Group")


def collect_message(update, context):
    message = update.message
    username = message.from_user.username
    chat_id = message.chat_id
    chat_type = message.chat.type
    text = message.text

    if chat_type == "private":
        if username not in ["Jellys04", "Cryptomaker143", "Shankar332", "Royce73", "Balaharishb", "LEO_sweet_67",
                            "SaranKMC", "pugalkmc", "SebastienKulec"]:
            bot.sendMessage(chat_id=chat_id, text="You have no permission to use this bot")
            return
        if "spreadsheet admin" == text:
            save_to_spreadsheet(update, context, admin="yes")
        elif "spreadsheet" in message.text and len(message.text) > 12:
            text = text.replace("spreadsheet ", "")
            save_to_spreadsheet(update=update, context=context, date=text)


    elif chat_type == "group" or chat_type == "supergroup":
        if chat_id not in [-1001588000922] or username not in ["Jellys04", "Cryptomaker143", "Shankar332", "Royce73",
                                                               "Balaharishb",
                                                               "LEO_sweet_67",
                                                               "SaranKMC", "pugalkmc"]:
            return

        # Only process messages from specific users in personal chat
        collection_name = datetime.now().strftime("%Y-%m-%d")
        message_id = message.message_id
        message_date_ist = (datetime.now() + timedelta(hours=5, minutes=30)).strftime(
            "%H:%M:%S")  # Convert datetime to IST timezone
        message_text = message.text

        # Store message data in Firebase Realtime Database
        db.reference(f'messages/{collection_name}/{message_id}').set({
            'username': username,
            'text': message_text,
            'time': message_date_ist,
            'message_id': message_id
        })


admins_list = [1155684571, 814546021, 1291659507]


def save_to_spreadsheet(update, context, admin=None, date=None):
    collection_name = date if date else datetime.now().strftime("%Y-%m-%d")
    # collection_name = (datetime.now() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d")

    # Get all the messages from the database for a specific date
    messages = db.reference(f'messages/{collection_name}').get() or {}

    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the headers
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws['A1'] = 'Username'
    ws['B1'] = 'Message Link'
    ws['C1'] = 'Message Text'
    ws['D1'] = 'IST Time'
    ws['F1'] = 'Username'
    ws['G1'] = 'Message Count'

    # Write the data
    row = 2
    username_counts = {}
    for message_id, message_data in messages.items():
        username = message_data.get('username')
        text = message_data.get('text')
        time = message_data.get('time')
        link = f'https://t.me/poolsea/{message_id}'

        if username:
            if username in username_counts:
                username_counts[username]['count'] += 1
            else:
                username_counts[username] = {'count': 1, 'total': 0}

            ws.cell(row=row, column=1).value = username
            ws.cell(row=row, column=2).value = link
            ws.cell(row=row, column=3).value = text
            ws.cell(row=row, column=4).value = time
            row += 1
    msg = ""
    for i in username_counts:
        msg = f"Username: {i}"
    bot.sendMessage(chat_id=update.message.chat_id, text=f"Total Messages: {len(messages.items())}\n\n"
                                                          f"{msg}")

    ws["F1"] = "Usernames"
    ws["G1"] = "Count"

    ws['F2'] = 'Jellys04'
    ws['F3'] = 'Cryptomaker143'
    ws['F4'] = 'Shankar332'
    ws['F5'] = "Royce73"
    ws['F6'] = "Balaharishb"
    ws['F7'] = "LEO_sweet_67"
    ws['F8'] = "SaranKMC"
    ws['F9'] = "pugalkmc"

    # set the formula in cell G2
    for row in range(2, 10):
        username = ws.cell(row=row, column=6).value  # Get the username from Column F
        count = '=COUNTIF(A:A,"*' + username + '*")'  # Construct the formula
        ws.cell(row=row, column=7).value = count

    # Save the Excel workbook
    file_name = f"{collection_name}.xlsx"
    wb.save(file_name)
    bot.sendDocument(chat_id=update.message.chat_id, document=open(file_name, 'rb'))
    if admin == "yes":
        for i in admins_list:
            bot.sendDocument(chat_id=i, document=open(file_name, "rb"))


def main():
    updater = Updater(token="6208523031:AAFfOb97T6Wml0pZUagE56A_MZDpCpUXZJk", use_context=True)
    dp = updater.dispatcher
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("spreadsheet", save_to_spreadsheet))
    dp.add_handler(MessageHandler(Filters.text, collect_message))
    updater.start_polling()


main()
